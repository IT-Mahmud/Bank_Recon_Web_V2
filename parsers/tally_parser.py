# tally_parser.py

import re
import string
import pandas as pd
from openpyxl import load_workbook
from calendar import month_name

# ---- BEGIN BANK MAPPING ----
BANK_ACCT_MAP = {
    # Mutual Trust Bank specific mappings
    "Mutual Trust Bank Ltd-SND-002-0320004355  Book":     ("MTB", "002-0320004355"),
    
    # One Bank specific mappings
    "One Bank-CD/A/C-0011020008826":                ("OBL", "0011020008826"),
    
    # Midland Bank specific mappings
    "Midland Bank PLC-CD-A/C-0011-1050011026  Book":    ("MDB", "0011-1050011026"),
    "Midland-CE-0011-1060000331-CI  Book":              ("MDB", "0011-1060000331"),
    "Midland-CE-0011-1060000322-CI  Book":              ("MDB", "0011-1060000322"),
    "Midland-CE-0011-1060000304-CI  Book":              ("MDB", "0011-1060000304"),

    # Eastern Bank specific mappings
    "Eastern Bank,STD-1011220144056":               ("EBL", "1011220144056"),
    "Eastern Bank Limited-SND-1011060605503":       ("EBL", "1011060605503"),
    
    # Prime Bank specific mappings
    "Prime Bank-CD-2126117010855":                  ("PBL", "2126117010855"),
    # Add more as needed...
}
# ---- END BANK MAPPING ----

def clean(val):
    return str(val).strip() if val is not None else ""

def deduplicate_row(row, dup_map):
    res = row[:]
    for val, idxs in dup_map.items():
        found = False
        for i in idxs:
            if clean(res[i]) == val:
                if found:
                    res[i] = ""
                else:
                    found = True
    return res

def process_particulars(value):
    if pd.isna(value):
        return ""
    val = str(value).replace('\r\n', '\n').replace('\r', '\n').strip()
    lines = [line.strip() for line in val.split('\n') if line.strip()]
    if not lines:
        return ""
    header = lines[0]
    if len(lines) > 1:
        details = ' '.join(lines[1:]).strip()
        return f"{header}\n{details}"
    match = re.match(r'^([A-Za-z0-9\-/ ]+)[.:,-]\s*(.+)', header)
    if match:
        header_part = match.group(1).strip()
        detail_part = match.group(2).strip()
        return f"{header_part}\n{detail_part}"
    return header

def extract_vendor_updated(particulars):
    if pd.isna(particulars):
        return ""
    val_str = str(particulars).strip()
    lines = [line.strip() for line in val_str.splitlines() if line.strip()]
    if lines and lines[0].lower().replace(" ", "") == "(asperdetails)" and len(lines) > 1:
        val = lines[1]
        match = re.search(r'\b([A-Za-z]+-CE-\d+-\d+-CI)\b', val)
        if match:
            return match.group(1).upper().replace(" ", "")
        match2 = re.search(r'Payable-([^-]+)-ID', val)
        if match2:
            return match2.group(1).strip().upper().replace(" ", "")
        match3 = re.search(r'([A-Za-z .&-]+(?:Ltd|Limited))', val, re.IGNORECASE)
        if match3:
            return match3.group(1).strip().upper().replace(" ", "")
        if 'Amount' in val:
            prefix = val.split('Amount')[0]
            chunks = [c.strip() for c in prefix.split('-') if c.strip()]
            if chunks:
                return re.sub(r'[^A-Z0-9]', '', chunks[-1].upper())
        return val.upper().replace(" ", "")
    else:
        val = lines[0] if lines else ""
        val = re.sub(r'^(adv(?:ance)?|ap)[\s\-]*', '', val, flags=re.IGNORECASE)
        val = re.sub(r'^(m[\s\-\/]*s)[\s\-]*', '', val, flags=re.IGNORECASE)
        val = re.split(r'-ID:', val, flags=re.IGNORECASE)[0]
        val = re.sub(r'\band\b', '', val, flags=re.IGNORECASE)
        val = re.sub(f'[{re.escape(string.punctuation)}\s]+', '', val)
        return val.upper()

def parse_tally_file(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # --------- Find header row ---------
    header_keywords = {"Date", "Particulars", "Vch Type", "Vch No.", "Debit", "Credit"}
    header_row_idx = next((i for i, r in enumerate(ws.iter_rows(values_only=True), 1)
                           if header_keywords.issubset({clean(c) for c in r})), None)
    if not header_row_idx:
        wb.close()
        raise ValueError("Header row not found.")

    # --------- Extract metadata for iloc use ---------
    # Build metadata as DataFrame for iloc reference
    metadata_rows = []
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx-1, values_only=True):
        metadata_rows.append([clean(c) for c in row])
    metadata = pd.DataFrame(metadata_rows)

    # --------- Hardcoded account extraction ---------
    try:
        acct_val = metadata.iloc[3, 0]
        acct_no_match = re.search(r'(\d{4}-\d{7,})', acct_val)
        acct_no = acct_no_match.group(1) if acct_no_match else acct_val
        bank_code = None

        for key, (code, ac_no) in BANK_ACCT_MAP.items():
            if key == acct_val:
                bank_code = code
                acct_no = ac_no
                break

        if not bank_code:
            bank_code = ""
    except Exception:
        acct_no = ""
        bank_code = ""

    if not bank_code or not acct_no:
        wb.close()
        raise ValueError(
            "Unmapped bank account detected or account cell missing. Update BANK_ACCT_MAP or check metadata."
        )

    # --------- Hardcoded statement period extraction ---------
    try:
        ledger_period_cell = metadata.iloc[6, 0]
    except Exception:
        wb.close()
        raise Exception("Could not access metadata cell [5, 0] for ledger period.")

    if not ledger_period_cell or "to" not in ledger_period_cell:
        wb.close()
        raise Exception("Ledger period cell missing or misformatted in [5, 0].")

    match = re.search(r'([\d]{1,2}-[A-Za-z]{3}-[\d]{4})\s*to\s*([\d]{1,2}-[A-Za-z]{3}-[\d]{4})', ledger_period_cell)
    if not match:
        wb.close()
        raise Exception("Could not parse dates in ledger period cell.")

    first_date_str, last_date_str = match.group(1), match.group(2)
    try:
        first_date = pd.to_datetime(first_date_str, format="%d-%b-%Y")
        last_date = pd.to_datetime(last_date_str, format="%d-%b-%Y")
    except Exception:
        wb.close()
        raise Exception("Date format error in ledger period cell.")

    ledger_date = ""
    ledger_year = ""
    if first_date.month == last_date.month:
        ledger_date = month_name[first_date.month]
    if first_date.year == last_date.year:
        ledger_year = str(first_date.year)

    # --------- Extract unit name from first cell (A1) ---------
    meta_unit_row = ws['A1'].value or ""
    unit_match = re.search(r'Unit\s*:?[\s)]*([^)]+)', meta_unit_row)
    unit_name = unit_match.group(1).strip() if unit_match else ""

    # --------- Unmerge all merged cells ---------
    for rng in list(ws.merged_cells.ranges):
        val = ws[rng.coord.split(":")[0]].value
        ws.unmerge_cells(str(rng))
        for row in ws[rng.coord]:
            for cell in row:
                cell.value = val

    headers = [clean(c.value) if c.value else f"Unnamed_{i+1}" for i, c in enumerate(ws[header_row_idx])]

    # Rename "Particulars" column to "dr_cr"
    headers = ["dr_cr" if h == "Particulars" and i == headers.index("Particulars") else h for i, h in enumerate(headers)]

    # Rename the next column to 'Particulars'
    particulars_index = headers.index("dr_cr") + 1  # Get the index of the next column after "dr_cr"
    if particulars_index < len(headers):
        headers[particulars_index] = "Particulars"

    num_cols = len(headers)

    collapsed_rows = []
    current_row = None
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cleaned = [clean(c) for c in row][:num_cols] + [""] * (num_cols - len(row))
        if (
            (not cleaned[headers.index("Date")] if "Date" in headers else True)
            and (not cleaned[headers.index("dr_cr")] if "dr_cr" in headers else True)
            and (cleaned[headers.index("Particulars")] if "Particulars" in headers else False)
            and current_row is not None
        ):
            idx = headers.index("Particulars")
            current_row[idx] = (current_row[idx] + "\n" + cleaned[idx]).strip()
        else:
            if current_row is not None:
                collapsed_rows.append(current_row)
            current_row = cleaned
    if current_row is not None:
        collapsed_rows.append(current_row)

    wb.close()
    data_rows = collapsed_rows
    dedup_map = {v: idxs for v, idxs in pd.Series(data_rows[0]).groupby(lambda x: x).groups.items() if len(idxs) > 1}
    data_rows = [deduplicate_row(row, dedup_map) for row in data_rows]

    if all(clean(v).replace('.', '', 1).replace(',', '', 1).isdigit() or clean(v) == "" for v in data_rows[-1]):
        data_rows.pop(-1)

    df = pd.DataFrame(data_rows, columns=headers).dropna(axis=1, how='all')
    df = df.loc[:, (df != '').any(axis=0)]
    df = df.loc[:, ~df.columns.str.match(r'Unnamed_\d+')]

    # Process "Particulars" column formatting and vendor extraction
    if "Particulars" in df.columns:
        df["Particulars"] = df["Particulars"].apply(process_particulars)
        df["tally_ven"] = df["Particulars"].apply(extract_vendor_updated)
    else:
        df["tally_ven"] = ""

    # Normalize date columns to yyyy-mm-dd format
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    # Remove rows where Particulars == 'Opening Balance' or starts with 'Closing Balance'
    if "Particulars" in df.columns:
        df = df[df["Particulars"].str.strip().str.lower() != "opening balance"]
        df = df[~df["Particulars"].str.strip().str.lower().str.startswith("closing balance")]

    def to_hex(val):
        try:
            return hex(int(float(val)))[2:]
        except Exception:
            return ""

    uids = []
    rownum = 1
    for i, row in df.iterrows():
        date_val = row.get("Date", "")
        credit_val = row.get("Credit", "")
        debit_val = row.get("Debit", "")
        balance_val = credit_val if (pd.notna(credit_val) and str(credit_val).strip() != "") else debit_val
        if pd.notna(date_val) and date_val != "":
            date_str = str(date_val).replace("-", "")
            hexdate = to_hex(date_str)
            try:
                hexbal = to_hex(round(float(str(balance_val).replace(",", ""))))
            except Exception:
                hexbal = ""
            # uid = f"T_{bank_code}_{rownum:04d}_{hexdate}_{hexbal}"
            uid = f"T_{bank_code}_{hexdate}_{hexbal}_{rownum:06d}"

            uids.append(uid)
            rownum += 1
        else:
            uids.append("")
    df["tally_uid"] = uids
    # cols = ["tally_uid", "bank_code", "bank_acct_no", "unit_name", "statement_month", "statement_year"] + [c for c in df.columns if c not in ["tally_uid", "bank_code", "bank_acct_no", "unit_name", "statement_month", "statement_year"]]
    cols = ["tally_uid", "bank_code", "acct_no", "unit_name", "statement_month", "statement_year"] + [c for c in df.columns if c not in ["tally_uid", "bank_code", "acct_no", "unit_name", "statement_month", "statement_year"]]

    df["bank_code"] = bank_code
    df["acct_no"] = acct_no.replace("-", "")  # Remove dashes from account number
    df["unit_name"] = unit_name
    df["statement_month"] = ledger_date
    df["statement_year"] = ledger_year
    df = df[cols]

    # Optionally: convert empty Debit/Credit to None for DB
    if "Debit" in df.columns:
        df["Debit"] = df["Debit"].apply(lambda x: None if str(x).strip() == '' else x)
    if "Credit" in df.columns:
        df["Credit"] = df["Credit"].apply(lambda x: None if str(x).strip() == '' else x)


    # Rename columns before writing to df
    new_column_names = {
        "Date": "T_Date",               # Renaming 'Date' to 'T_Date'
        "Particulars": "T_Particulars",  # Newly renamed Particulars to T_Particulars
        "Vch Type": "T_Vch_Type",        # Renaming 'Vch Type' to 'T_Vch_Type'
        "Vch No.": "T_Vch_No",           # Renaming 'Vch No.' to 'T_Vch_No'
        "Debit": "T_Debit",               # Renaming 'Debit' to 'T_Debit'
        "Credit": "T_Credit",             # Renaming 'Credit' to 'T_Credit'
        # Add any other columns you want to rename here
    }

    # Apply the column renaming to df
    df = df.rename(columns=new_column_names)

    # Return the final df with the renamed columns
    return df  